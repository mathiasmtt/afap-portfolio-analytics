"""Generación de reportes Excel multi-hoja para el equipo comercial.

El dashboard Streamlit expone un botón "Descargar Excel" que invoca
:func:`generar_excel` y devuelve los bytes del libro en memoria.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.pipeline import Pipeline

from src.analytics import (
    aplicar_segmentaciones,
    cohortes_afiliacion,
    heatmap_edad_antiguedad,
    kpis_globales,
    pareto_cartera,
    tasa_fuga_por_segmento,
    variacion_entre_segmentos,
)
from src.models.churn_logit import drivers_principales, scorear_afiliados

HEADER_FILL = PatternFill(start_color="FFEC7000", end_color="FFEC7000", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)


def _escribir_hoja(wb: Workbook, nombre: str, df: pd.DataFrame) -> None:
    """Escribe un DataFrame en una hoja con formato corporativo."""
    ws = wb.create_sheet(title=nombre[:31])  # Excel: máx 31 chars
    if df.empty:
        ws.append(list(df.columns) if len(df.columns) else ["(sin datos)"])
        return

    ws.append(list(df.columns))
    for celda in ws[1]:
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila in df.itertuples(index=False, name=None):
        ws.append(list(fila))

    # Anchos automáticos estimados.
    for idx, col in enumerate(df.columns, start=1):
        muestra = df[col].astype(str).head(50)
        ancho = min(max(len(str(col)), *(len(v) for v in muestra)) + 2, 30)
        ws.column_dimensions[get_column_letter(idx)].width = ancho


def _hoja_resumen(kpis: dict[str, float]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"KPI": "N° afiliados", "Valor": kpis["n_afiliados"]},
            {"KPI": "Saldo total (UYU)", "Valor": round(kpis["saldo_total"], 2)},
            {"KPI": "Ticket promedio (UYU)", "Valor": round(kpis["ticket_promedio"], 2)},
            {"KPI": "Tasa de fuga", "Valor": round(kpis["tasa_fuga"], 4)},
            {"KPI": "% aportantes activos", "Valor": round(kpis["pct_activos"], 4)},
        ]
    )


def generar_excel(
    df: pd.DataFrame,
    model: Pipeline | None = None,
    top_riesgo_n: int = 100,
) -> bytes:
    """Genera un libro Excel multi-hoja con los insights del dashboard.

    Parameters
    ----------
    df : DataFrame AFAP completo.
    model : pipeline entrenado (opcional). Si se provee, agrega la hoja
        'Top Riesgo' con el scoring de los ``top_riesgo_n`` afiliados.
    top_riesgo_n : cantidad de afiliados en la hoja de scoring.

    Returns
    -------
    Bytes del archivo ``.xlsx`` listos para `st.download_button`.
    """
    wb = Workbook()
    # Remover la hoja default creada por openpyxl.
    default = wb.active
    wb.remove(default)

    kpis = kpis_globales(df)
    _escribir_hoja(wb, "Resumen", _hoja_resumen(kpis))

    _escribir_hoja(wb, "Pareto Cartera", pareto_cartera(df).head(500))

    seg = aplicar_segmentaciones(df)
    fuga_depto = tasa_fuga_por_segmento(df, "departamento")
    fuga_edad = tasa_fuga_por_segmento(seg, "tramo_edad")
    fuga_antig = tasa_fuga_por_segmento(seg, "tramo_antiguedad")
    fuga_saldo = tasa_fuga_por_segmento(seg, "tramo_saldo")
    segmentacion = pd.concat(
        [
            fuga_depto.assign(dimension="departamento").rename(columns={"departamento": "valor"}),
            fuga_edad.assign(dimension="tramo_edad").rename(columns={"tramo_edad": "valor"}),
            fuga_antig.assign(dimension="tramo_antiguedad").rename(
                columns={"tramo_antiguedad": "valor"}
            ),
            fuga_saldo.assign(dimension="tramo_saldo").rename(columns={"tramo_saldo": "valor"}),
        ],
        ignore_index=True,
    )[["dimension", "valor", "n_afiliados", "n_fugas", "tasa_fuga", "saldo_total"]]
    _escribir_hoja(wb, "Segmentacion", segmentacion)

    _escribir_hoja(wb, "Cohortes", cohortes_afiliacion(df))

    desvios = variacion_entre_segmentos(df, "departamento")
    _escribir_hoja(wb, "Desvios", desvios)

    heat = heatmap_edad_antiguedad(df).reset_index().rename(columns={"index": "tramo_edad"})
    _escribir_hoja(wb, "Heatmap Edad x Antig", heat)

    if model is not None:
        scores = scorear_afiliados(model, df).head(top_riesgo_n)
        # Enriquecer con saldo y departamento para que el equipo llame hoy.
        top = scores.merge(
            df[["afiliado_id", "apellido", "departamento", "edad", "saldo_cuenta"]],
            on="afiliado_id",
            how="left",
        )
        _escribir_hoja(wb, "Top Riesgo", top)
        _escribir_hoja(wb, "Drivers Modelo", drivers_principales(model))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
