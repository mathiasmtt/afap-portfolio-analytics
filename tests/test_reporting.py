"""Tests para src.reporting."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.models.churn_logit import entrenar_logit
from src.reporting import generar_excel


@pytest.fixture(scope="module")
def excel_bytes_sin_modelo(df_bancario_sintetico: pd.DataFrame) -> bytes:
    from src.data_loader import reframe_bancario_a_afap

    df = reframe_bancario_a_afap(df_bancario_sintetico)
    return generar_excel(df)


@pytest.fixture(scope="module")
def excel_bytes_con_modelo(df_bancario_sintetico: pd.DataFrame) -> bytes:
    from src.data_loader import reframe_bancario_a_afap

    df = reframe_bancario_a_afap(df_bancario_sintetico)
    modelo = entrenar_logit(df, seed=42)
    return generar_excel(df, modelo, top_riesgo_n=50)


class TestExcelSinModelo:
    def test_bytes_no_vacio(self, excel_bytes_sin_modelo: bytes) -> None:
        assert len(excel_bytes_sin_modelo) > 1000

    def test_hojas_esperadas(self, excel_bytes_sin_modelo: bytes) -> None:
        wb = load_workbook(BytesIO(excel_bytes_sin_modelo))
        esperadas = {"Resumen", "Pareto Cartera", "Segmentacion", "Cohortes", "Desvios"}
        assert esperadas.issubset(set(wb.sheetnames))

    def test_resumen_tiene_datos(self, excel_bytes_sin_modelo: bytes) -> None:
        wb = load_workbook(BytesIO(excel_bytes_sin_modelo))
        ws = wb["Resumen"]
        assert ws.max_row >= 2  # header + al menos 1 fila

    def test_no_incluye_top_riesgo_sin_modelo(self, excel_bytes_sin_modelo: bytes) -> None:
        wb = load_workbook(BytesIO(excel_bytes_sin_modelo))
        assert "Top Riesgo" not in wb.sheetnames


class TestExcelConModelo:
    def test_incluye_hojas_de_modelo(self, excel_bytes_con_modelo: bytes) -> None:
        wb = load_workbook(BytesIO(excel_bytes_con_modelo))
        assert "Top Riesgo" in wb.sheetnames
        assert "Drivers Modelo" in wb.sheetnames

    def test_top_riesgo_respeta_n(self, excel_bytes_con_modelo: bytes) -> None:
        wb = load_workbook(BytesIO(excel_bytes_con_modelo))
        ws = wb["Top Riesgo"]
        # header + 50 filas (top_riesgo_n=50)
        assert ws.max_row == 51
